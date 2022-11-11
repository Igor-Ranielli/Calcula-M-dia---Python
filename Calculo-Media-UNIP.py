import PySimpleGUI as sg
import openpyxl

class TelaPython:
    def __init__ (self):
        sg.theme('Black')
        layout = [
        [sg.Text('Informe as notas de prova, pim e ava.')],
        [sg.Text('Nome da Prova'), sg.Input(size = (19,0),key='nome')],
        [sg.Text('Prova',size = (23,0)), sg.Input(size = (5,0),key='prova')],
        [sg.Text('Pim',size = (23,0)), sg.Input(size = (5,0),key='pim')],
        [sg.Text('Ava',size = (23,0)), sg.Input(size = (5,0),key='ava')],
        [sg.Button('Calcular a Media',size = (29,0))],
        [sg.Output(size=(32,10))]
       ]

        self.janela = sg.Window("Calculo de Media").layout(layout)
    def Iniciar(self):
        try:
            book = openpyxl.load_workbook('Planilha de Provas.xlsx')
        except:
            book = openpyxl.Workbook()
            book.create_sheet('Provas', 0)
        while True:
            self.button, self.values = self.janela.Read()
            if self.button == sg.WINDOW_CLOSED:
                break
            else:
                nome = self.values['nome']
                prova = self.values['prova']
                pim = self.values['pim']
                ava = self.values['ava']
                while True:
                        try:
                            prova = float(self.values['prova'])
                            pim = float(self.values['pim'])
                            ava = float(self.values['ava'])
                        except:
                            [sg.popup(f'ERRO! Digite um número válido.')]
                            break
                        if (prova > 10) or (pim > 10) or (ava > 10):
                            [sg.popup(f'ERRO! Digite uma nota entre 0 e 10')]
                            break
                        elif (prova < 0) or (pim < 0) or (ava < 0):
                            [sg.popup(f'ERRO! Digite uma nota entre 0 e 10')]
                            break
                        else: 
                            provas_page = book['Provas']
                            provas_page.column_dimensions['A'].width = 15
                            provas_page.column_dimensions['B'].width = 50
                            provas_page.column_dimensions['C'].width = 8
                            provas_page.column_dimensions['D'].width = 5
                            media = prova*7 + pim*2 + ava
                            notaFinal= media/10
                            print(f'Média final: {notaFinal}')
                            provas_page.append(['Nome da Prova:', nome, 'Média:', notaFinal])
                            book.save('Planilha de Provas.xlsx')
                            break

tela = TelaPython()
tela.Iniciar()
